import openpyxl
from os import path

def cargar_archivo( ruta_archivo ):
    if ( path.exists( ruta_archivo ) ):
        # Retorna un archivo en excel que ya se encuentra creado.
        return openpyxl.load_workbook( ruta_archivo )
    else:
        # Retorna un archivo nuevo.
        return openpyxl.Workbook()


def obtener_hoja( objeto, titulo_hoja ):
    if (objeto.sheetnames[0] != titulo_hoja):
        # El archivo es nuevo.
        hoja = objeto['Sheet']
        hoja.title = titulo_hoja
    else:
        # El archivo ya existe y no tiene la hoja "Sheet"
        hoja = objeto[ titulo_hoja ]
        
    return hoja
