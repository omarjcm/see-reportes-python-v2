import openpyxl
from os import path
from datetime import datetime

def obtener_nombre_archivo():
    '''
        %Y -> Año
        %b -> Mes
        %d -> Día
        %H -> Hora
        %M -> Minutos
        %S -> Segundos
    '''
    fecha_hora = datetime.now()
    fecha_hora_formato = fecha_hora.strftime( '%Y-%b-%d_%H-%M-%S' )
    return '../Data/reporte_see_'+fecha_hora_formato+'.xlsx'


def cargar_archivo( ruta_archivo ):
    if ( path.exists( ruta_archivo ) ):
        # Retorna un archivo en excel que ya se encuentra creado.
        return openpyxl.load_workbook( ruta_archivo )
    else:
        # Retorna un archivo nuevo.
        return openpyxl.Workbook()


def obtener_hoja( objeto, titulo_hoja ):
    if (objeto.sheetnames[0] != titulo_hoja):
        # El archivo es nuevo.
        hoja = objeto['Sheet']
        hoja.title = titulo_hoja
    else:
        # El archivo ya existe y no tiene la hoja "Sheet"
        hoja = objeto[ titulo_hoja ]
        
    return hoja
