import pandas
from datos import *
from openpyxl.styles import numbers

def agregar_datos( ws, sheet, cabecera ):
    df_paises, df_transacciones = get_transacciones_x_pais()

    inicial = 8
    final = len( df_paises ) + inicial
    sheet.insert_rows(inicial, final)
    #paises = df_paises.columns
        
    for index_row, data_row in df_paises.iterrows():
        index_col = 1
        data = list( data_row )
            
        sheet.cell(row=inicial, column=index_col, value=data[0])
        index_col+=1
        sheet.cell(row=inicial, column=index_col, value=data[1])
        index_col+=1
                
        cabecera_index = 2
        while cabecera_index < len(cabecera):
            # Comparo país, año y mes para agregar celda por celda.
            loc_country = df_transacciones['Country'] == data[0]
            loc_anio = df_transacciones['anio'] == data[1]
            loc_mes = df_transacciones['mes'] == int(cabecera[cabecera_index])
            
            ventas_x_mes = df_transacciones.loc[ loc_country & loc_anio & loc_mes ]
            
            celda = None
            if len(ventas_x_mes) > 0:
                celda = sheet.cell(row=inicial, column=index_col)
                celda.value = ventas_x_mes['venta_usd'].values[0]
            else:
                celda = sheet.cell(row=inicial, column=index_col, value='')
            celda.number_format = numbers.FORMAT_CURRENCY_USD
            index_col += 1
            cabecera_index += 1
            
        inicial+=1