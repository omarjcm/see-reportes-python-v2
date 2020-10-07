import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

borde_fino = Side( border_style='thin', color= '0089bb')
borde_titulo = Border( top=borde_fino, bottom=borde_fino, left=borde_fino, right=borde_fino )
fondo_titulo = PatternFill( 'solid', fgColor='0089bb'  )
fuente_titulo = Font( name='Calibri', size=18, b=True, color='ffffff' )
alineacion_titulo = Alignment( horizontal='center', vertical='center' )
fuente_subtitulo = Font( name='Calibri', size=14, b=True, color='ffffff' )

def agregar_imagen( hoja ):
    imagen = openpyxl.drawing.image.Image( '../Data/images/see_webinar.png' )
    hoja.add_image( imagen, 'A1' )


def agregar_titulo( hoja ):
    hoja.oddHeader.left.color = 'CC3366'
    hoja.merge_cells( 'A5:N5' )
    celda = hoja['A5']
    celda.value = 'Reporte de los Países'
    
    celda.border = borde_titulo
    celda.fill = fondo_titulo
    celda.font = fuente_titulo
    celda.alignment = alineacion_titulo


def agregar_filtro( hoja_activa, hoja, cabecera ):
    celdas = hoja['A7:N7']
    
    indice = 0
    for columna in celdas:
        for celda in columna:
            celda.value = cabecera[indice]
            celda.fill = fondo_titulo
            celda.font = fuente_subtitulo
            celda.alignment = alineacion_titulo
            indice += 1
        
    hoja_activa.auto_filter.ref = 'A7:N7'
            
    








